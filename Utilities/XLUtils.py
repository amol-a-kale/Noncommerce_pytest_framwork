import openpyxl


def getrowcount(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return (sheet.max_row)


def getcolumncount(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return (sheet.max_column)


def readData(filename, sheetname, rownum, columnum):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=columnum).value


def writeData(filename,sheetName,rownum,columnum,data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnum).value = data
    workbook.save(filename)


